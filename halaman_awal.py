import csv
from tkinter import ttk
from pathlib import Path
import customtkinter as ctk
from PIL import Image, ImageTk
import openpyxl

OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path(r"D:\SELF ORDER MACHINE\assets")

def menu_list():
    window4 = ctk.CTkToplevel()
    window4.title("WARPAT!")
    window4.geometry('1366x768')
    window4.state('zoomed')
    window4.configure(bg="#fbcd64")
    window4.resizable(True, True)

    bg_welcome = ImageTk.PhotoImage(Image.open("assets/background_menu.png"))
    bgr = ctk.CTkLabel(master=window4, image=bg_welcome, text="")
    bgr.pack()

    def button_dine(path: str) -> Path:
        return ASSETS_PATH / Path(path)

    button_makanan = ctk.CTkImage(Image.open(button_dine("button_makan.png")), size=(322.38, 61.85))
    button_4 = ctk.CTkButton(
        window4,
        image=button_makanan,
        border_width=0,
        corner_radius=0,
        bg_color="#fbcd64",
        fg_color="#fbcd64",
        hover_color="#fbcd64",
        command=lambda: [window4.destroy(), list_makanan()],
        text=''
    )
    button_4.place(x=-30, y=190.16)

    button_minuman = ctk.CTkImage(Image.open(button_dine("button_minum.png")), size=(322.38, 61.85))
    button_5 = ctk.CTkButton(
        window4,
        image=button_minuman,
        border_width=0,
        corner_radius=0,
        bg_color="#fbcd64",
        fg_color="#fbcd64",
        hover_color="#fbcd64",
        command=lambda: [window4.destroy(), list_minuman()],
        text=''
    )
    button_5.place(x=-30, y=264.9)

    button_extra = ctk.CTkImage(Image.open(button_dine("button_extra.png")), size=(322.38, 61.85))
    button_6 = ctk.CTkButton(
        window4,
        image=button_extra,
        border_width=0,
        corner_radius=0,
        bg_color="#fbcd64",
        fg_color="#fbcd64",
        hover_color="#fbcd64",
        command=lambda: [window4.destroy(), list_extra()],
        text=''
    )
    button_6.place(x=-30, y=339.16)

def list_minuman():
    window5 = ctk.CTkToplevel()
    window5.title("WARPAT!")
    window5.geometry('1366x768')
    window5.state('zoomed')
    window5.configure(bg="#fbcd64")
    window5.resizable(True, True)

    def baca_csv(file_path):
        data = []
        with open(file_path, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                data.append(row)
        return data
    
    def buat_label(frame, text, row, col):
        label = ctk.CTkLabel(master=frame, text=text)
        label.grid(row=row, column=col, padx=10, pady=5, sticky="w")
        return label
    
    main_frame = ctk.CTkFrame(master=window5)
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)

    canvas = ctk.CTkCanvas(master=main_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(master=main_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = ctk.CTkFrame(master=canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

    canvas.configure(yscrollcommand=scrollbar.set)

    data_makanan = baca_csv("menu_minum.csv")

    for i, item in enumerate(data_makanan):
        buat_label(scrollable_frame, item["Menu"], i, 0)
        buat_label(scrollable_frame, item["Harga"], i, 1)

    window5.mainloop()

def list_makanan():
    window6 = ctk.CTkToplevel()
    window6.title("WARPAT!")
    window6.geometry('1366x768')
    window6.state('zoomed')
    window6.configure(bg="#fbcd64")
    window6.resizable(True, True)

    def baca_csv(file_path):
        data = []
        with open(file_path, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                data.append(row)
        return data
    
    def buat_label(frame, text, row, col):
        label = ctk.CTkLabel(master=frame, text=text)
        label.grid(row=row, column=col, padx=10, pady=5, sticky="w")
        return label
    
    main_frame = ctk.CTkFrame(master=window6)
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)

    canvas = ctk.CTkCanvas(master=main_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(master=main_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = ctk.CTkFrame(master=canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

    canvas.configure(yscrollcommand=scrollbar.set)

    data_makanan = baca_csv("menu_makan.csv")

    for i, item in enumerate(data_makanan):
        buat_label(scrollable_frame, item["Menu"], i, 0)
        buat_label(scrollable_frame, item["Harga"], i, 1)

    window6.mainloop()

def list_extra():
    window7 = ctk.CTkToplevel()
    window7.title("WARPAT!")
    window7.geometry('1366x768')
    window7.state('zoomed')
    window7.configure(bg="#fbcd64")
    window7.resizable(True, True)

    def baca_csv(file_path):
        data = []
        with open(file_path, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                data.append(row)
        return data
    
    def buat_label(frame, text, row, col):
        label = ctk.CTkLabel(master=frame, text=text)
        label.grid(row=row, column=col, padx=10, pady=5, sticky="w")
        return label
    
    main_frame = ctk.CTkFrame(master=window7)
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)

    canvas = ctk.CTkCanvas(master=main_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(master=main_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = ctk.CTkFrame(master=canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

    canvas.configure(yscrollcommand=scrollbar.set)

    data_makanan = baca_csv("menu_makan.csv")

    for i, item in enumerate(data_makanan):
        buat_label(scrollable_frame, item["Menu"], i, 0)
        buat_label(scrollable_frame, item["Harga"], i, 1)

    window7.mainloop()

def dine_option():
    window3 = ctk.CTkToplevel()
    window3.title("WARPAT!")
    window3.geometry('1366x768')
    window3.state('zoomed')
    window3.configure(bg="#fbcd64")
    window3.resizable(True, True)

    bg_welcome = ImageTk.PhotoImage(Image.open("assets/background_saji.png"))
    bgr = ctk.CTkLabel(master=window3, image=bg_welcome, text="")
    bgr.pack()

    def button_dine(path: str) -> Path:
        return ASSETS_PATH / Path(path)

    button_image_3 = ctk.CTkImage(Image.open(button_dine("button_makansini.png")), size=(307.36, 214.87))
    button_3 = ctk.CTkButton(
        window3,
        image=button_image_3,
        border_width=0,
        corner_radius=0,
        bg_color="#fbcd64",
        fg_color="#fbcd64",
        hover_color="#fbcd64",
        command=lambda: [window3.destroy(), menu_list()],
        text=''
    )
    button_3.place(x=310.0, y=250.0)

    button_image_4 = ctk.CTkImage(Image.open(button_dine("button_bungkus.png")), size=(307.36, 214.87))
    button_4 = ctk.CTkButton(
        window3,
        image=button_image_4,
        border_width=0,
        corner_radius=0,
        bg_color="#fbcd64",
        fg_color="#fbcd64",
        hover_color="#fbcd64",
        command=lambda: [window3.destroy(), menu_list()],
        text=''
    )
    button_4.place(x=650.0, y=250.0)

    window3.mainloop()

def halaman_nama():
    def save_to_excel(nama):
        file_path = "id_nama.xlsx"
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet['A1'] = 'Nama'

        next_row = sheet.max_row + 1
        sheet[f'A{next_row}'] = nama

        workbook.save(file_path)

    def on_button_click():
        nama = entry_nama.get()
        if nama:  
            save_to_excel(nama)
            window2.destroy()
            dine_option()  

    window2 = ctk.CTkToplevel()
    window2.title("WARPAT!")
    window2.geometry('1366x768')
    window2.state('zoomed')
    window2.configure(bg="#fbcd64")
    window2.resizable(True, True)

    bg_welcome = ImageTk.PhotoImage(Image.open("assets/background_nama.png"))
    bgr = ctk.CTkLabel(master=window2, image=bg_welcome, text="")
    bgr.pack()

    entry_nama = ctk.CTkEntry(
        window2,
        placeholder_text="Masukkan Nama Anda",
        height=58.1,
        width=791,
        font=("Helvetica", 18),
        corner_radius=15,
        text_color='#FFFFFF',
        placeholder_text_color="#C7411E",
        fg_color=("#F28743"),
        bg_color="#fbcd64",
        border_width=3,
        border_color=('#C7411E')
    )
    entry_nama.place(x=240, y=0)

    def button_nama(path: str) -> Path:
        return ASSETS_PATH / Path(path)

    button_masuk_img = ctk.CTkImage(Image.open(button_nama("button_masuk.png")), size=(147.29, 53.97))
    button_2 = ctk.CTkButton(
        window2,
        image=button_masuk_img,
        width=147.29,
        height=53.97,
        border_width=0,
        corner_radius=0,
        bg_color="#48775D",
        fg_color="#48775D",
        hover_color="#48775D",
        command=on_button_click,
        text=''
    )
    button_2.place(x=1100.0, y=665.0)

    window2.mainloop()

def halaman_awal():
    window = ctk.CTk()
    window.title("WARPAT!")
    window.geometry('1920x1200')
    window.state('zoomed')
    window.configure(bg="#fbcd64")
    window.resizable(True, True)

    bg_welcome = ImageTk.PhotoImage(Image.open("assets/background_welcome.png"))
    bgr = ctk.CTkLabel(master=window, image=bg_welcome, text="", width=1366, height=768)
    bgr.pack()

    def button_welcome(path: str) -> Path:
        return ASSETS_PATH / Path(path)

    button_start_image = ctk.CTkImage(Image.open(button_welcome("button_logo.png")), size=(250, 250))
    button_1 = ctk.CTkButton(
        window,
        image=button_start_image,
        width=200,
        height=200,
        corner_radius=0,
        bg_color="#fbcd64",
        fg_color="transparent",
        hover_color="#fbcd64",
        command=lambda: [window.destroy(), halaman_nama()],
        border_width=0,
        text=''
    )
    button_1.place(x=515.0, y=230.0)

    window.mainloop()

halaman_awal()
